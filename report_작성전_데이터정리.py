from openpyxl import load_workbook
import datetime
import os
# 화면 지우고 프로그램 시작하기
os.system('cls' if os.name == 'nt' else 'clear')

# 프로그램에 대한 정보
print('''
보고서 전처리 자동화 프로그램

작업할 보고서는 [oa report 2023 06.xlsx] 형식으로 저장하시기 바랍니다.

작업 결과물은 [oa report 2023 06 output.xlsx] 형식으로 같은 폴더에 저장됩니다.
''')

# 몇 월의 보고서인지 입력 받기 / 숫자가 아니면 재입력 오류
while True:  
    # 월 입력 받기
    month_input = input("몇 월의 보고서를 불러오시겠습니까? (숫자로 입력) : ")
    try:
        month = int(month_input)
        break  # 정수 값이 입력되면 빠져나가기
    except ValueError:
        print("\n잘못 입력 하셨네요. 숫자를 입력해주세요.\n")

# 엑셀 파일 경로 생성
name = "oa report 2023 {:02d}".format(month)
file_name = name+".xlsx"

# 엑셀 파일 불러오기
workbook = load_workbook(file_name, read_only=False)
sheet = workbook.active

# 변경할 시트 이름들
target_sheets1_1 = ["업무실적"]

# target_sheet1_1에 대한 작업
for sheet_name in target_sheets1_1:
    sheet = workbook[sheet_name]
    # 데이터 이동
    for row in range(24, 30):
        value_i = sheet.cell(row=row, column=9).value  # I열의 값
        sheet.cell(row=row, column=7).value = value_i  # G열로 이동
        sheet.cell(row=row, column=9).value = None  # 원래 위치의 값을 삭제

        value_j = sheet.cell(row=row, column=10).value  # J열의 값
        sheet.cell(row=row, column=8).value = value_j  # H열로 이동
        sheet.cell(row=row, column=10).value = None  # 원래 위치의 값을 삭제


target_sheets1 = ["지역별업무실적"]
for sheet_name in target_sheets1:
    sheet = workbook[sheet_name]
    
    for row in range(4, 9):
        value = sheet.cell(row=row, column=3).value
        sheet.cell(row=row, column=2).value = value
        sheet.cell(row=row, column=3).value = None
        
    for row in range(4, 9):
        value = sheet.cell(row=row, column=5).value
        sheet.cell(row=row, column=4).value = value
        sheet.cell(row=row, column=5).value = None
        
    for row in range(4, 9):
        value = sheet.cell(row=row, column=7).value
        sheet.cell(row=row, column=6).value = value
        sheet.cell(row=row, column=7).value = None
        
    for row in range(4, 9):
        value = sheet.cell(row=row, column=9).value
        sheet.cell(row=row, column=8).value = value
        sheet.cell(row=row, column=9).value = None
        
    for row in range(4, 9):
        value = sheet.cell(row=row, column=11).value
        sheet.cell(row=row, column=10).value = value
        sheet.cell(row=row, column=11).value = None
        
        
target_sheets2 = ["장애HW", "장애SW", "통신업무", "장비업무", "지원업무", "관리업무"]
for sheet_name in target_sheets2:
    sheet = workbook[sheet_name]
    
    for row in range(2, 33):
        sheet.row_dimensions[row].height = 18
    has_data = False
    
    for row in range(3, 21):
        value = sheet.cell(row=row, column=3).value
        sheet.cell(row=row, column=2).value = value
        sheet.cell(row=row, column=3).value = None
        
    for row in range(3, 21):
        sheet.cell(row=row, column=5).value = None


target_sheets1_2 = ["표지"]
for sheet_name in target_sheets1_2:
    sheet = workbook[sheet_name]

    month = datetime.datetime.now().month
    sheet['A5'] = month
    
    today = datetime.datetime.now().strftime("%Y-%m-%d")
    sheet['C8'] = today

# 변경된 내용 저장 (기존 파일 덮어쓰기)
new_file_name = name+" output.xlsx"
workbook.save(new_file_name)

# 현재 실행 중인 스크립트 파일 경로
script_path = os.path.abspath(__file__)

# 실행 중인 스크립트 파일의 폴더 경로
folder_path = os.path.dirname(script_path)

# 탐색기에서 실행 경로 폴더 열기
os.system(f'explorer "{folder_path}"')
